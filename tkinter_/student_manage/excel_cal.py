from openpyxl import load_workbook

import pandas as pd
import numpy as np

# 计算总分并返回学号和姓名、学号和总分的字典
# id2deduct_score：key:学号，value:扣分
def cal_total_score(data, ws):
    id2deduct_score = {}
    id2name = {}
    col_start, col_end = cal_deduct_score_range(data)
    row_start = 1; row_end = 0
    for index, row in data.iterrows():
        # 过滤掉不是学号的列
        if not isinstance(row.values[0], int):
            continue
        # 过滤掉没有考试的学生
        cell_bg_color = ws.cell(row=index+2, column=1).fill.start_color.index
        if cell_bg_color != '00000000' and cell_bg_color != 0:
            continue
        def is_plus_number(value):
            return isinstance(value, str) and value.startswith('+')
        # 统计每一个学号对应的分数
        # 将带加号的字符串转换为NaN
        row_cleaned = row[col_start:col_end+1].apply(lambda x: np.nan if is_plus_number(x) else x)
        deduct_points = pd.to_numeric(row_cleaned, errors='coerce').sum(skipna=True)
        id2deduct_score[row.values[0]] = deduct_points
        id2name[row.values[0]] = row.values[1]
        row_end = index
    return id2deduct_score, id2name, row_start, row_end, col_start, col_end

# 按照总分排序，并返回学号和排名的字典
# id2sort：分数排名：key:学号，value:排名位置
def sort_total_score(id2deduct_score):
    id2sort = {}
    # 按照分数降序排列
    sorted_students = sorted(id2deduct_score.items(), key=lambda item: item[1], reverse=False)
    last_score = None
    last_sort = 0
    for i, (student, score) in enumerate(sorted_students):
        if i==0 or score != last_score:
            last_score = score
            last_sort = i+1
        id2sort[student] = last_sort
    return id2sort

# 将总分和排名写回到表格中
def write_total_score(id2deduct_score, id2name, id2sort, data, filename, wb, ws, row_start, row_end, col_start, col_end):
    # 遍历每一行并在指定列写入数据
    deduct_index, id_index = cal_write_index(data)
    id_sort = list(id2sort.items())
    no_examination_count = 0
    for index, row in data.iterrows():
        c_id = row.values[0]
        # 过滤掉不是学号的列 和 没有分数的行
        if not isinstance(row.values[0], int):
            continue
        #  扣分项的单独统计，如果没有成绩则不参与统计
        if id2deduct_score.get(c_id) is None:
            no_examination_count += 1
        else:
            # 假设我们将结果写入 'B' 列（Excel中的第2列）
            ws.cell(row=index + 2, column=deduct_index, value=id2deduct_score[c_id])
            ws.cell(row=index + 2, column=deduct_index + 1, value=100-id2deduct_score[c_id])  # +2是因为 DataFrame 索引从 0 开始，Excel 从 1 开始，且第一行通常是表头
            ws.cell(row=index + 2, column=deduct_index + 2, value=id2sort[c_id])

        # 后面顺序统计项，不空行连续展示
        # 统计表中获取本行的学号
        index = index - no_examination_count
        s_id, sort = id_sort[index-1]
        ws.cell(row=index + 2, column=id_index, value=s_id)
        ws.cell(row=index + 2, column=id_index + 1, value=id2name[s_id])
        ws.cell(row=index + 2, column=id_index + 2, value=100-id2deduct_score[s_id])
        ws.cell(row=index + 2, column=id_index + 3, value=sort)

    # 计算得分率
    for col_index, c in enumerate(data.columns):
        if col_index < col_start or col_index > col_end:
            continue
        try:
            # 获取每一题的分数
            single_score = data.iloc[row_end + 1, col_index]
            if pd.isna(single_score):
                continue
            dect_total = data.iloc[row_start:row_end+1, col_index].sum()
            percentage = '%.2f'%((single_score * len(id2name) - dect_total) / (single_score * len(id2name)))
            ws.cell(row=row_end + 2 + 1, column=col_index + 1, value=percentage)
        except Exception as e:
            pass

    # 计算班级均分
    avg_score = '%.2f'%((len(id2deduct_score)*100 - sum(id2deduct_score.values())) / len(id2deduct_score))
    ws.cell(row=row_end + 2 + 2, column=col_end + 1 + 3, value=avg_score)
    # 保存工作簿
    wb.save(filename)

# 检查文件找出统计扣分的列范围
def cal_deduct_score_range(data):
    start = 0
    end = 0
    title_row = data.iloc[0]
    for col_index, value in enumerate(title_row):
        if value == '姓名':
            start = col_index + 1
        if value == '扣分':
            end = col_index - 1
            break
    return start, end

# 查出排名统计的列范围
def cal_write_index(data):
    deduct_index = 0
    id_index = 0
    title_row = data.iloc[0]
    for col_index, value in enumerate(title_row):
        if value == '扣分':
            deduct_index = col_index
        if value == '学号':
            id_index = col_index
    return deduct_index + 1, id_index + 1
